#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Mon Oct 23 16:37:03 2017

@author: apple
"""
from openpyxl import load_workbook
# Load file and read the data
def read_the_data(path):
    workBook = load_workbook(path)                 # Load the file
    dataSheet = workBook.get_sheet_by_name('TrainData')  # Get the datasheet

# Range from the datasheet to assign the training set, test set and validation set.
    row = 378742
    col = 5
    dataSet = [[0] * col] * row
    for i in range(0, row):
        for j in range(0, col):
            dataSet[i][j] = dataSheet.cell(row = i + 1, column = j + 1).value
    return dataSet