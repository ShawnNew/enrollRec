#%% Recommendation system for students applying for university
# Author: Chenxiao Niu
# Date: Oct, 24,2017
# Version: 1.0.0
# Platform: TensorFlow
# @CopyRight

#%% Import Section
import tensorflow as tf 
from openpyxl import load_workbook
import numpy as np
import math

#%% Define next_batch function
def next_batch(num, data, labels):
    '''
    Return a total of 'num' random samples and labels.
    '''
    if len(data) == len(labels):
        idx = np.arange(0, len(data))
        np.random.shuffle(idx)
        idx = idx[:num]
        data_shuffle = [data[i] for i in idx]
        labels_shuffle = [labels[i] for i in idx]
        return np.asarray(data_shuffle), np.asarray(labels_shuffle)
    else:
        print 'The data set does not match with labels'
        
#%% Define normalization function
def normalize_columns(arr):
    rows, cols = arr.shape
    for col in xrange(cols):
        arr[:,col] /= abs(arr[:,col]).max()
    return arr

#%% Define weights
def init_weights(shape):
    return tf.Variable(tf.random_normal(shape, stddev = 0.01))

#%% Define model
def model(X, w_layer_1, w_layer_2, w_layer_3, w_layer_4, w_layer_5, p_keep_input, p_keep_hidden):
    X = tf.nn.dropout(X, p_keep_input)
    hidden_1 = tf.nn.relu(tf.matmul(X, w_layer_1))

    hidden_1 = tf.nn.dropout(hidden_1, p_keep_hidden)
    hidden_2 = tf.nn.relu(tf.matmul(hidden_1, w_layer_2))

    hidden_2 = tf.nn.dropout(hidden_2, p_keep_hidden)
    hidden_3 = tf.nn.relu(tf.matmul(hidden_2, w_layer_3))
    
    hidden_3 = tf.nn.dropout(hidden_3, p_keep_hidden)
    hidden_4 = tf.nn.relu(tf.matmul(hidden_3, w_layer_4))
    
    hidden_4 = tf.nn.dropout(hidden_4, p_keep_hidden)
    

    return tf.matmul(hidden_4, w_layer_5)


#%% Load the data from xlsx file
print('Loading the file...')
path = "/Users/apple/Documents/Github/enrollRec/stuData.xlsx"
workBook = load_workbook(path)                 # Load the file
dataSheet = workBook.get_sheet_by_name('TrainData')  # Get the datasheet
print 'Loading file Done...'

#%% From the file read the dataSet
print('Reading the file...')
row = 378742
col = 6
dataSet = np.zeros((row, col))


for i in range(0, row):
    for j in range(0, col):
        dataSet[i][j] = dataSheet.cell(row = i + 1, column = j + 1).value
    #label[i] = dataSheet.cell(row = i + 1, column = 6).value
print('Reading file done...')
   
     
#%% Divide the dataSet into training set, test set and validation set
print 'Construction the data...'

numTrainSet = int(math.ceil(row * 0.7))
numTestSet = int(math.ceil(row * 0.15)) - 1
numValSet = int(math.ceil(row * 0.15)) - 1
#data_label_combined = np.hstack((dataSet, label))  # Combine the dataset and label to shuffle

np.random.shuffle(dataSet)             # Shuffle the dataset to divide

# Divide the dataset
data_train_withlabel = dataSet[0: numTrainSet, :]
data_test_withlabel = dataSet[numTrainSet: (numTrainSet + numTestSet), :]
data_val_withlabel = dataSet[(numTrainSet + numTestSet) : (numTrainSet + numTestSet + numValSet), :]
# Normalize the dataSet
data_train = normalize_columns(data_train_withlabel[:, 0 : 5])
data_val = normalize_columns(data_val_withlabel[:, 0: 5])
data_test = normalize_columns(data_test_withlabel[:, 0: 5])
# Vectorize the label
label_train = np.zeros((numTrainSet, 10))
label_test = np.zeros((numTestSet, 10))
label_val = np.zeros((numValSet, 10))
for i in range(len(data_train_withlabel)):
    for j in range(0, 10):
        label_train[i][(int(data_train_withlabel[i][5])) - 1] = 1.0
for i in range(len(data_test_withlabel)):
    for j in range(0, 10):
        label_test[i][(int(data_test_withlabel[i][5])) - 1] = 1.0
for i in range(len(data_val_withlabel)):
    for j in range(0, 10):
        label_val[i][(int(data_val_withlabel[i][5])) - 1] = 1.0

print 'Constructing the data done...'
print 'Size of Training Set:' + str(len(data_train))
print 'Size of Test Set:' + str(len(data_test))
print 'Size of Validation Set:' + str(len(data_val))

#%% Create the Graph
# dataset and label
print 'Creating Network...'
X = tf.placeholder(tf.float32, [None, 5])
Y = tf.placeholder(tf.float32, [None, 10])

# Define the paras of the layers
# The weights
n_input = 5   # 5 features
n_hidden_1 = 100
n_hidden_2 = 100
n_hidden_3 = 100
n_hidden_4 = 100
n_classes = 10


# Input layer to hidden_layer 1
w_layer_1 = init_weights([n_input, n_hidden_1])
# Hidden_layer_1 to hidden_layer 2
w_layer_2 = init_weights([n_hidden_1, n_hidden_2])
# Hidden_layer_2 to hidden_layer 3
w_layer_3 = init_weights([n_hidden_2, n_hidden_3])
# Hidden_layer 3 to hidden_layer 4
w_layer_4 = init_weights([n_hidden_3, n_hidden_4])
# Hidden_layer 4 to output
w_layer_5 = init_weights([n_hidden_4, n_classes])

# dropout
# Dropout rate for input layer
p_keep_input = tf.placeholder(tf.float32)
# Dropout rate for hidden layer
p_keep_hidden = tf.placeholder(tf.float32)

# Construct the model
py_x = model(X, w_layer_1, w_layer_2, w_layer_3, w_layer_4, w_layer_5, p_keep_input, p_keep_hidden)

cost = tf.reduce_mean(tf.nn.softmax_cross_entropy_with_logits(logits=py_x, labels=Y))
train_op = tf.train.RMSPropOptimizer(1e-2, 0.3).minimize(cost)
#train_op = tf.train.GradientDescentOptimizer(learning_rate = 1e-3).minimize(cost)
predict_op = tf.argmax(py_x, 1)
print 'Network Ready...'



#keep_prob = tf.placeholder(tf.float32)
#with tf.name_scope("Weight1") as scope:
#    W1 = tf.Variable(tf.random_normal([n_input, n_hidden_1], -1.0, 1.0))
#    #W1_hist = tf.summary.histogram("Weight1", W1)
#with tf.name_scope("Weight2") as scope:
#    W2 = tf.Variable(tf.random_normal([n_hidden_1, n_hidden_2], -1.0, 1.0))
#    #W2_hist = tf.summary.histogram("Weight2", W2)
#with tf.name_scope("Weight3") as scope:
#    W3 = tf.Variable(tf.random_normal([n_hidden_2, n_hidden_3], -1.0, 1.0))
#    #W3_hist = tf.summary.histogram("Output", W3)
#with tf.name_scope("Weight4") as scope:
#    W4 = tf.Variable(tf.random_normal([n_hidden_3, n_hidden_4], -1.0, 1.0))
#with tf.name_scope("Weight5") as scope:
#    W5 = tf.Variable(tf.random_normal([n_hidden_4, n_classes], -1.0, 1.0))
#
## Bias
#b1 = tf.Variable(tf.random_normal([n_hidden_1]), name = "Bias1")
#b2 = tf.Variable(tf.random_normal([n_hidden_2]), name = "Bias2")
#b3 = tf.Variable(tf.random_normal([n_hidden_3]), name = "Bias3")
#b4 = tf.Variable(tf.random_normal([n_hidden_4]), name = "Bias4")
#b5 = tf.Variable(tf.random_normal([n_classes]), name = "Bias5")
#print 'Network Ready...'
#
## Hypothesis
#with tf.name_scope("input") as scope:
#    L1 = tf.nn.relu(tf.matmul(X, W1) + b1)
#    L1 = tf.nn.dropout(L1, keep_prob=keep_prob)  #dropout to prevent overfitting
#with tf.name_scope("layer2") as scope:
#    L2 = tf.nn.relu(tf.matmul(L1, W2) + b2)
#    L2 = tf.nn.dropout(L2, keep_prob=keep_prob)
#with tf.name_scope("layer3") as scope:
#    L3 = tf.nn.relu(tf.matmul(L2, W3) + b3)
#    L3 = tf.nn.dropout(L3, keep_prob=keep_prob)
#with tf.name_scope("layer4") as scope:
#    #hypothesis = tf.nn.sigmoid(tf.matmul(L2, W3) + b3)
#    L4 = tf.nn.relu(tf.matmul(L3, W4) + b4)
#    L4 = tf.nn.dropout(L4, keep_prob=keep_prob)
#with tf.name_scope("output") as scope:
#    hypothesis = tf.nn.sigmoid(tf.matmul(L4, W5) + b5)
#
#    
## cost/loss function
##cost = tf.reduce_mean(-tf.reduce_sum(Y * tf.log(hypothesis), reduction_indices=[1]))
#
#cost = - tf.reduce_mean(Y * tf.log(tf.clip_by_value(hypothesis, 1e-10,tf.reduce_max(hypothesis))) +
#                       (1 - Y) * tf.log(tf.clip_by_value((1 - hypothesis), 1e-10,tf.reduce_max(hypothesis))))
##cost = tf.reduce_mean(tf.nn.softmax(hypothesis, Y, name='cost')) 
#train = tf.train.GradientDescentOptimizer(learning_rate = 10).minimize(cost)
#prediction = tf.argmax(hypothesis, 1)
#label = tf.argmax(Y, 1)
#correct_prediction = tf.equal(prediction, label)
#accuracy = tf.reduce_mean(tf.cast(correct_prediction, tf.float32))


#%% Run the graph
# Initialize TensorFlow variablies
training_epochs = 200000
batch_size      = 100
display_step    = 200



with tf.Session() as sess:

    init = tf.global_variables_initializer()
    sess.run(init)

    for i in xrange(training_epochs):
        batch_xs, batch_ys = next_batch(batch_size, data_train, label_train)
        sess.run(train_op, feed_dict = {X: batch_xs, Y: batch_ys,
                                            p_keep_input: 0.8, p_keep_hidden: 0.5})
        if i % display_step == 0: 
            cost_at_running = sess.run(cost, feed_dict = {X: batch_xs, Y: batch_ys, p_keep_input: 1.0, p_keep_hidden: 1.0})
            print 'Batch ' + str(i) + '\ncost is: ' + str(cost_at_running) + '        Accuracy:    ' + str(np.mean(np.argmax(batch_ys, axis = 1) == sess.run(predict_op, 
                        feed_dict = {X: batch_xs, Y: batch_ys, p_keep_input: 1.0, p_keep_hidden: 1.0})))
           




#sess = tf.InteractiveSession()
#tf.global_variables_initializer().run()
#for i in range(training_epochs):
## Iterate 10000 epoches, and every epoch with a batch of 100 data
#    batch_xs, batch_ys = next_batch(batch_size, data_train, label_train)
#    sess.run(train, feed_dict = {X: batch_xs, Y: batch_ys, keep_prob: 0.7})
#    # hy = sess.run(hypothesis, feed_dict={X: batch_xs, Y: batch_ys, keep_prob: 0.7})
#    # print hy.shape
#    if i % display_step == 0:
#        cost_at_running, acc = sess.run([cost, accuracy], feed_dict={X: batch_xs, Y: batch_ys, keep_prob:0.7})
#        target, pred = sess.run([label, prediction], feed_dict={X:batch_xs, Y:batch_ys, keep_prob:1})
#        print 'target is:   ' + str(target) + '         prediction is:   ' + str(pred)
#        print 'Batch ' + str(i) + '\ncost is: ' + str(cost_at_running) + '       Current accuracy is: ' + str(acc)
        

#%% Test model and check accuracy

# correct_prediction = tf.equal(tf.argmax(hypothesis, 1), tf.argmax(label_train, 1))
# accuracy = tf.reduce_mean(tf.cast(accuracy, tf.float32))




#acc = sess.run(accuracy, feed_dict={X: data_test, Y: label_test, keep_prob: 1})
#print('Accuracy:' + str(acc))
    
    
    
    


